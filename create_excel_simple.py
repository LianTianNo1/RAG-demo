#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
简单的Excel文件创建脚本

@remarks 使用openpyxl直接创建Excel文件，不依赖pandas
@author AI Assistant
@version 1.0
"""

import os

def create_simple_excel():
    """
    使用openpyxl创建简单的Excel文件
    
    @returns Excel文件路径或None
    @example
    ```python
    excel_path = create_simple_excel()
    if excel_path:
        print(f"Excel文件创建成功: {excel_path}")
    ```
    """
    try:
        from openpyxl import Workbook
        
        # 确保data目录存在
        data_dir = "./data/"
        if not os.path.exists(data_dir):
            os.makedirs(data_dir)
            print(f"创建了数据目录: {data_dir}")

        # 创建工作簿
        wb = Workbook()
        
        # 删除默认工作表
        wb.remove(wb.active)
        
        # 创建员工信息工作表
        ws1 = wb.create_sheet("员工信息")
        ws1.append(['姓名', '部门', '职位', '入职年份', '薪资(万)'])
        ws1.append(['张三', '技术部', '软件工程师', 2020, 15])
        ws1.append(['李四', '市场部', '市场专员', 2021, 12])
        ws1.append(['王五', '技术部', '项目经理', 2019, 18])
        ws1.append(['Alice', '人事部', 'HR专员', 2022, 10])
        
        # 创建项目信息工作表
        ws2 = wb.create_sheet("项目信息")
        ws2.append(['项目名称', '负责人', '状态', '预算(万)', '开始时间'])
        ws2.append(['Alpha项目', '王五', '进行中', 50, '2023-01-15'])
        ws2.append(['Beta项目', '张三', '已完成', 30, '2022-06-01'])
        ws2.append(['Gamma项目', '李四', '规划中', 70, '2023-03-01'])
        
        # 创建部门统计工作表
        ws3 = wb.create_sheet("部门统计")
        ws3.append(['部门', '人数', '平均薪资(万)', '部门预算(万)'])
        ws3.append(['技术部', 2, 16.5, 100])
        ws3.append(['市场部', 1, 12, 50])
        ws3.append(['人事部', 1, 10, 30])
        
        # 保存文件
        excel_file_path = os.path.join(data_dir, "sample_data.xlsx")
        wb.save(excel_file_path)
        
        print(f"成功创建示例Excel文件: {excel_file_path}")
        print("文件包含以下工作表:")
        print("  - 员工信息: 包含员工的基本信息")
        print("  - 项目信息: 包含项目的详细信息")
        print("  - 部门统计: 包含各部门的统计数据")
        
        return excel_file_path
        
    except ImportError:
        print("openpyxl未安装，请运行: pip install openpyxl")
        return None
    except Exception as e:
        print(f"创建Excel文件时发生错误: {e}")
        return None

if __name__ == "__main__":
    create_simple_excel()
